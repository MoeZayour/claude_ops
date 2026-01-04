
import base64
import json
import openpyxl
import xlsxwriter
from io import BytesIO
from odoo import models, fields, api, _
from odoo.exceptions import UserError

class SaleOrderImportWizard(models.TransientModel):
    _name = 'sale.order.import.wizard'
    _description = 'Sale Order Import from Excel Wizard'

    import_file = fields.Binary(
        string='Excel File',
        required=True,
        help='Upload your Excel file with columns: Section, Model, Quantity'
    )
    import_filename = fields.Char(
        string='Filename'
    )
    template_file = fields.Binary(
        string='Template',
        compute='_compute_template_file',
        readonly=True
    )
    import_result = fields.Selection([
        ('pending', 'Pending'),
        ('success', 'Success'),
        ('error', 'Error'),
    ], default='pending', string='Import Result')
    error_message = fields.Text(string='Error Messages')
    success_message = fields.Text(string='Success Summary')
    lines_imported = fields.Integer(string='Lines Imported', default=0)
    validated_rows = fields.Text(string='Validated Rows (JSON)', help='Stores validated data for Session 3')

    def _compute_template_file(self):
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('Import Template')

        # Write headers
        headers = ['Section', 'Model', 'Quantity']
        for i, header in enumerate(headers):
            worksheet.write(0, i, header)

        # Write sample data
        sample_data = [
            ('Electronics', 'iPhone 15 Pro', 5),
            ('Electronics', 'iPad Air', 3),
            ('Accessories', 'Apple Pencil', 8),
            ('Office Supplies', 'Notebook A4', 50),
        ]
        for r, row_data in enumerate(sample_data, 1):
            for c, cell_data in enumerate(row_data):
                worksheet.write(r, c, cell_data)
        
        workbook.close()
        output.seek(0)
        self.template_file = base64.b64encode(output.read())

    def action_download_template(self):
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content?model=sale.order.import.wizard&id={self.id}&field=template_file&download=true&filename=SO_Import_Template.xlsx',
            'target': 'self',
        }

    def action_validate_and_import(self):
        # 1. Read file
        if not self.import_file:
            raise UserError('Please upload an Excel file')
        
        # Check file size (5 MB = 5 * 1024 * 1024 bytes)
        file_size = len(base64.b64decode(self.import_file))
        max_size = 5 * 1024 * 1024

        if file_size > max_size:
            raise UserError(f'File size ({file_size / (1024*1024):.1f} MB) exceeds 5 MB limit.')

        file_data = base64.b64decode(self.import_file)
        workbook = openpyxl.load_workbook(BytesIO(file_data))
        sheet = workbook.active
        
        # 2. Validate file structure
        structure_errors = self._validate_file_structure(sheet)
        if structure_errors:
            return self._show_error_wizard(structure_errors)
        
        # 3. Validate all rows
        errors = []
        validated_rows = []
        
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue
                
            section = row[0]
            model = row[1]
            quantity = row[2]
            
            # Validate this row
            row_errors = self._validate_row(idx, section, model, quantity)
            
            if row_errors:
                errors.extend(row_errors)
            else:
                validated_rows.append({
                    'section': section,
                    'model': model,
                    'quantity': quantity,
                })
        
        # 4. Check for errors (all-or-nothing)
        if errors:
            return self._show_error_wizard(errors)
        
        # 5. Success - store validated data for Session 3
        self.write({
            'validated_rows': json.dumps(validated_rows),
            'lines_imported': len(validated_rows),
        })
        
        return self._show_success_wizard(len(validated_rows))

    def _validate_file_structure(self, sheet):
        errors = []
        
        # Check minimum rows
        if sheet.max_row < 2:
            errors.append('File is empty or contains no data rows')
            return errors
        
        # Check max rows (1,000 + 1 header = 1,001 max)
        if sheet.max_row > 1001:
            errors.append(f'File contains {sheet.max_row - 1} rows. Maximum allowed is 1,000.')
            return errors
        
        # Check columns (exactly 3)
        if sheet.max_column != 3:
            errors.append(f'Invalid column structure. Expected 3 columns, found {sheet.max_column}.')
            return errors
        
        # Check headers
        headers = [cell.value for cell in sheet[1]]
        expected = ['Section', 'Model', 'Quantity']
        
        for i, expected_header in enumerate(expected):
            if not headers[i] or headers[i].strip().lower() != expected_header.lower():
                errors.append(f'Invalid header in column {i+1}. Expected "{expected_header}", found "{headers[i]}"')
        
        return errors

    def _validate_row(self, row_num, section, model, quantity):
        errors = []
        
        # Required fields
        if not section:
            errors.append(f'Row {row_num}: Section is required')
        if not model:
            errors.append(f'Row {row_num}: Model is required')
        if not quantity:
            errors.append(f'Row {row_num}: Quantity is required')
            return errors  # Can't validate further without quantity
        
        # Product exists (search by name OR default_code, case-insensitive)
        product = self.env['product.product'].search([
            '|',
            ('name', '=ilike', model),
            ('default_code', '=ilike', model)
        ], limit=1)
        
        if not product:
            errors.append(f'Row {row_num}: Product \'{model}\' not found. Check spelling or use Internal Reference.')
        
        # Quantity validation
        try:
            qty = float(quantity)
            if qty <= 0:
                errors.append(f'Row {row_num}: Quantity must be greater than 0')
        except (ValueError, TypeError):
            errors.append(f'Row {row_num}: Quantity \'{quantity}\' is not a valid number')
        
        return errors

    def _show_error_wizard(self, errors):
        error_text = '\n'.join([f'❌ {error}' for error in errors])
        
        self.write({
            'import_result': 'error',
            'error_message': f'Found {len(errors)} error(s) in your file:\n\n{error_text}\n\nNo lines were imported. Please fix the errors and try again.',
        })
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'sale.order.import.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': {'form_view_initial_mode': 'edit'},
        }

    def _show_success_wizard(self, count):
        """Show success wizard and create lines."""
        # Parse validated data
        validated_data = json.loads(self.validated_rows)
        
        # CREATE THE LINES (this is new!)
        total_imported = self._create_order_lines(validated_data)
        
        # Group by section for summary
        sections = {}
        for row in validated_data:
            section = row['section']
            sections[section] = sections.get(section, 0) + 1
        
        summary_lines = [f'📦 {section}: {count} products' for section, count in sections.items()]
        summary = '\n'.join(summary_lines)
        
        self.write({
            'import_result': 'success',
            'success_message': f'Successfully imported {total_imported} lines:\n\n{summary}\n\nLines have been added to your sales order.',
            'lines_imported': total_imported,
        })
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'sale.order.import.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': {'form_view_initial_mode': 'edit'},
        }

    def _create_order_lines(self, validated_rows):
        """Create sale order lines from validated data.
        
        Process:
        1. Get the sale order from context
        2. Group rows by section
        3. For each section:
           - Create section header (display_type='line_section')
           - Create product lines under that section
        4. Post message to sale order chatter
        """
        # Get sale order from context
        sale_order_id = self.env.context.get('active_id')
        if not sale_order_id:
            raise UserError('No sale order found. Please open this wizard from a sale order.')
        
        sale_order = self.env['sale.order'].browse(sale_order_id)
        
        # Group by section (maintain order)
        sections = {}
        section_order = []
        for row in validated_rows:
            section = row['section']
            if section not in sections:
                sections[section] = []
                section_order.append(section)
            sections[section].append(row)
        
        # Track statistics
        total_lines = 0
        
        # Create lines for each section
        for section_name in section_order:
            section_rows = sections[section_name]
            
            # Create section header
            sale_order.order_line.create({
                'order_id': sale_order.id,
                'display_type': 'line_section',
                'name': section_name.upper(),
                'sequence': 10000,  # Append to end
            })
            
            # Create product lines
            for row in section_rows:
                # Find product (we already validated it exists)
                product = self.env['product.product'].search([
                    '|',
                    ('name', '=ilike', row['model']),
                    ('default_code', '=ilike', row['model'])
                ], limit=1)
                
                if product:
                    # Create sale order line
                    sale_order.order_line.create({
                        'order_id': sale_order.id,
                        'product_id': product.id,
                        'product_uom_qty': float(row['quantity']),
                        'sequence': 10000,  # Append to end
                    })
                    total_lines += 1
        
        # Post to chatter
        sale_order.message_post(
            body=_(
                '<p><strong>Excel Import Completed</strong></p>'
                '<ul>'
                '<li>Total lines imported: <strong>%s</strong></li>'
                '<li>Sections: <strong>%s</strong></li>'
                '</ul>'
            ) % (total_lines, ', '.join(section_order)),
            message_type='notification',
            subtype_xmlid='mail.mt_note'
        )
        
        return total_lines
