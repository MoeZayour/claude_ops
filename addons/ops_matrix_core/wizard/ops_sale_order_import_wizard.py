# -*- coding: utf-8 -*-
from odoo import models, fields, api
from odoo.exceptions import UserError, ValidationError
import base64
import io
import logging

_logger = logging.getLogger(__name__)


class OpsSaleOrderImportWizard(models.TransientModel):
    _name = 'ops.sale.order.import.wizard'
    _description = 'Import Sale Order Lines from Excel'
    
    sale_order_id = fields.Many2one('sale.order', 'Sale Order', required=True)
    import_file = fields.Binary('Excel File', required=True)
    import_filename = fields.Char('Filename')
    template_file = fields.Binary('Template File', readonly=True, compute='_compute_template_file')
    template_filename = fields.Char('Template Filename', readonly=True, default='sale_order_import_template.xlsx')
    
    state = fields.Selection([
        ('upload', 'Upload File'),
        ('validate', 'Validation Results'),
        ('done', 'Import Complete'),
    ], default='upload', string='State')
    
    validation_message = fields.Html('Validation Results', readonly=True)
    lines_imported = fields.Integer('Lines Imported', readonly=True)
    
    @api.depends('sale_order_id')
    def _compute_template_file(self):
        """Generate Excel template with instructions"""
        for wizard in self:
            try:
                wizard.template_file = base64.b64encode(wizard._generate_template())
            except Exception as e:
                _logger.error(f"Error generating template: {e}")
                wizard.template_file = False
    
    def _generate_template(self):
        """Generate Excel template with instructions"""
        try:
            import xlsxwriter
        except ImportError:
            raise UserError("xlsxwriter Python library is not installed. Please contact your system administrator.")
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        
        # Instructions sheet
        instructions = workbook.add_worksheet('Instructions')
        bold = workbook.add_format({'bold': True, 'font_size': 12, 'font_color': '#1F4E78'})
        regular = workbook.add_format({'font_size': 10})
        
        instructions.write('A1', 'HOW TO USE THIS TEMPLATE', bold)
        instructions.write('A3', '1. Fill in the "Import Data" sheet with your products', regular)
        instructions.write('A4', '2. Product Code: Must match exactly an existing product code (case-sensitive)', regular)
        instructions.write('A5', '3. Quantity: Must be a positive number', regular)
        instructions.write('A6', '4. Unit Price: Optional - leave blank to use the product default price', regular)
        instructions.write('A7', '5. Discount: Optional - enter percentage (0-100)', regular)
        instructions.write('A8', '6. Save the file and upload it in Odoo', regular)
        instructions.write('A9', '7. Click "Validate File" to check for errors before importing', regular)
        instructions.write('A10', '8. Do NOT modify column headers!', bold)
        instructions.set_column('A:A', 70)
        
        # Data sheet with headers
        data_sheet = workbook.add_worksheet('Import Data')
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#366092',
            'font_color': 'white',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        headers = ['Product Code', 'Product Name (Reference Only)', 'Quantity', 'Unit Price (Optional)', 'Discount % (Optional)']
        for col, header in enumerate(headers):
            data_sheet.write(0, col, header, header_format)
        
        # Example data
        example_format = workbook.add_format({'italic': True, 'font_color': '#666666'})
        data_sheet.write(1, 0, 'PROD001', example_format)
        data_sheet.write(1, 1, 'Example Product Name', example_format)
        data_sheet.write(1, 2, 10, example_format)
        data_sheet.write(1, 3, 100.00, example_format)
        data_sheet.write(1, 4, 0, example_format)
        
        data_sheet.set_column('A:A', 20)
        data_sheet.set_column('B:B', 35)
        data_sheet.set_column('C:C', 12)
        data_sheet.set_column('D:D', 20)
        data_sheet.set_column('E:E', 20)
        
        workbook.close()
        return output.getvalue()
    
    def action_download_template(self):
        """Download template file"""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content?model={self._name}&id={self.id}&field=template_file&filename={self.template_filename}&download=true',
            'target': 'self',
        }
    
    def action_validate_import(self):
        """Validate Excel file"""
        self.ensure_one()
        
        if not self.import_file:
            raise UserError("Please upload an Excel file first.")
        
        errors = []
        warnings = []
        valid_lines = []
        
        try:
            # Try xlrd first for .xls files
            try:
                import xlrd
                file_data = base64.b64decode(self.import_file)
                workbook = xlrd.open_workbook(file_contents=file_data)
                use_xlrd = True
            except (ImportError, xlrd.biffh.XLRDError):
                use_xlrd = False
            
            # Fall back to openpyxl for .xlsx files
            if not use_xlrd:
                try:
                    from openpyxl import load_workbook
                    file_data = base64.b64decode(self.import_file)
                    workbook = load_workbook(io.BytesIO(file_data), read_only=True)
                    use_openpyxl = True
                except ImportError:
                    raise UserError("Neither xlrd nor openpyxl library is installed. Please contact your system administrator.")
            else:
                use_openpyxl = False
            
            # Check if "Import Data" sheet exists
            if use_xlrd:
                if 'Import Data' not in workbook.sheet_names():
                    raise UserError("Excel file must contain 'Import Data' sheet. Please use the template.")
                sheet = workbook.sheet_by_name('Import Data')
                
                # Validate headers
                expected_headers = ['Product Code', 'Product Name (Reference Only)', 'Quantity', 'Unit Price (Optional)', 'Discount % (Optional)']
                actual_headers = [sheet.cell_value(0, col).strip() if col < sheet.ncols else '' for col in range(5)]
                
                if actual_headers != expected_headers:
                    raise UserError(f"Column headers don't match template. Expected: {expected_headers}")
                
                # Process rows (skip header)
                for row_idx in range(1, sheet.nrows):
                    # Skip empty rows
                    if not sheet.cell_value(row_idx, 0):
                        continue
                    
                    row_num = row_idx + 1
                    product_code = str(sheet.cell_value(row_idx, 0)).strip()
                    quantity = sheet.cell_value(row_idx, 2) if sheet.ncols > 2 else 0
                    unit_price = sheet.cell_value(row_idx, 3) if sheet.ncols > 3 and sheet.cell_value(row_idx, 3) else None
                    discount = sheet.cell_value(row_idx, 4) if sheet.ncols > 4 and sheet.cell_value(row_idx, 4) else 0
                    
                    # Validate product exists
                    product = self.env['product.product'].search([
                        ('default_code', '=', product_code)
                    ], limit=1)
                    
                    if not product:
                        errors.append(f"Row {row_num}: Product code '{product_code}' not found in system")
                        continue
                    
                    # Validate quantity
                    try:
                        qty = float(quantity)
                        if qty <= 0:
                            errors.append(f"Row {row_num}: Quantity must be positive (got {qty})")
                            continue
                    except (ValueError, TypeError):
                        errors.append(f"Row {row_num}: Invalid quantity '{quantity}'")
                        continue
                    
                    # Validate price
                    if unit_price is not None:
                        try:
                            price = float(unit_price)
                            if price < 0:
                                errors.append(f"Row {row_num}: Price cannot be negative")
                                continue
                        except (ValueError, TypeError):
                            warnings.append(f"Row {row_num}: Invalid price, using default price for {product.name}")
                            price = product.list_price
                    else:
                        price = product.list_price
                    
                    # Validate discount
                    try:
                        disc = float(discount) if discount else 0
                        if disc < 0 or disc > 100:
                            warnings.append(f"Row {row_num}: Discount must be 0-100%, got {disc}%. Setting to 0.")
                            disc = 0
                    except (ValueError, TypeError):
                        disc = 0
                    
                    # Valid line
                    valid_lines.append({
                        'product_id': product.id,
                        'product_uom_qty': qty,
                        'price_unit': price,
                        'discount': disc,
                    })
            
            elif use_openpyxl:
                if 'Import Data' not in workbook.sheetnames:
                    raise UserError("Excel file must contain 'Import Data' sheet. Please use the template.")
                sheet = workbook['Import Data']
                
                # Validate headers
                expected_headers = ['Product Code', 'Product Name (Reference Only)', 'Quantity', 'Unit Price (Optional)', 'Discount % (Optional)']
                actual_headers = [str(sheet.cell(1, col+1).value or '').strip() for col in range(5)]
                
                if actual_headers != expected_headers:
                    raise UserError(f"Column headers don't match template. Expected: {expected_headers}")
                
                # Process rows (skip header row 1)
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    # Skip empty rows
                    if not row[0]:
                        continue
                    
                    product_code = str(row[0]).strip()
                    quantity = row[2] if len(row) > 2 and row[2] else 0
                    unit_price = row[3] if len(row) > 3 and row[3] else None
                    discount = row[4] if len(row) > 4 and row[4] else 0
                    
                    # Validate product exists
                    product = self.env['product.product'].search([
                        ('default_code', '=', product_code)
                    ], limit=1)
                    
                    if not product:
                        errors.append(f"Row {row_idx}: Product code '{product_code}' not found in system")
                        continue
                    
                    # Validate quantity
                    try:
                        qty = float(quantity)
                        if qty <= 0:
                            errors.append(f"Row {row_idx}: Quantity must be positive (got {qty})")
                            continue
                    except (ValueError, TypeError):
                        errors.append(f"Row {row_idx}: Invalid quantity '{quantity}'")
                        continue
                    
                    # Validate price
                    if unit_price is not None:
                        try:
                            price = float(unit_price)
                            if price < 0:
                                errors.append(f"Row {row_idx}: Price cannot be negative")
                                continue
                        except (ValueError, TypeError):
                            warnings.append(f"Row {row_idx}: Invalid price, using default price for {product.name}")
                            price = product.list_price
                    else:
                        price = product.list_price
                    
                    # Validate discount
                    try:
                        disc = float(discount) if discount else 0
                        if disc < 0 or disc > 100:
                            warnings.append(f"Row {row_idx}: Discount must be 0-100%, got {disc}%. Setting to 0.")
                            disc = 0
                    except (ValueError, TypeError):
                        disc = 0
                    
                    # Valid line
                    valid_lines.append({
                        'product_id': product.id,
                        'product_uom_qty': qty,
                        'price_unit': price,
                        'discount': disc,
                    })
            
            # Generate validation message
            msg = "<div style='padding: 15px; font-family: Arial, sans-serif;'>"
            
            if errors:
                msg += "<div style='background-color: #f8d7da; border-left: 4px solid #d9534f; padding: 10px; margin-bottom: 15px;'>"
                msg += "<h3 style='color: #d9534f; margin-top: 0;'>❌ Validation Failed</h3>"
                msg += "<p style='margin-bottom: 10px;'><strong>Please fix the following errors:</strong></p><ul style='margin: 0;'>"
                for error in errors:
                    msg += f"<li style='color: #721c24;'>{error}</li>"
                msg += "</ul></div>"
            
            if warnings:
                msg += "<div style='background-color: #fff3cd; border-left: 4px solid #f0ad4e; padding: 10px; margin-bottom: 15px;'>"
                msg += "<h3 style='color: #856404; margin-top: 0;'>⚠️ Warnings</h3><ul style='margin: 0;'>"
                for warning in warnings:
                    msg += f"<li style='color: #856404;'>{warning}</li>"
                msg += "</ul></div>"
            
            if valid_lines and not errors:
                msg += "<div style='background-color: #d4edda; border-left: 4px solid #5cb85c; padding: 10px;'>"
                msg += "<h3 style='color: #5cb85c; margin-top: 0;'>✅ Validation Successful</h3>"
                msg += f"<p style='color: #155724; margin: 0;'>Ready to import <strong>{len(valid_lines)} line(s)</strong> to sale order <strong>{self.sale_order_id.name}</strong></p>"
                msg += "<p style='color: #155724; margin-top: 10px;'>Click '<strong>Import Lines</strong>' below to add them to your sale order.</p>"
                msg += "</div>"
                
                # Store valid lines in config parameter (temporary storage)
                import json
                self.env['ir.config_parameter'].sudo().set_param(
                    f'ops_import_lines_{self.id}',
                    json.dumps(valid_lines)
                )
            
            msg += "</div>"
            
            self.write({
                'validation_message': msg,
                'state': 'validate' if not errors else 'upload',
            })
            
        except Exception as e:
            _logger.exception("Error processing Excel file")
            raise UserError(f"Error processing file: {str(e)}\n\nPlease ensure you're using the correct template and the file is not corrupted.")
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
    
    def action_import_lines(self):
        """Import validated lines"""
        self.ensure_one()
        
        # Retrieve validated lines
        import json
        lines_str = self.env['ir.config_parameter'].sudo().get_param(f'ops_import_lines_{self.id}')
        if not lines_str:
            raise UserError("No validated lines found. Please validate the file first.")
        
        valid_lines = json.loads(lines_str)
        
        # Import lines
        SaleOrderLine = self.env['sale.order.line']
        imported = 0
        
        for line_data in valid_lines:
            SaleOrderLine.create({
                'order_id': self.sale_order_id.id,
                'product_id': line_data['product_id'],
                'product_uom_qty': line_data['product_uom_qty'],
                'price_unit': line_data['price_unit'],
                'discount': line_data['discount'],
            })
            imported += 1
        
        # Clean up temp data
        self.env['ir.config_parameter'].sudo().set_param(f'ops_import_lines_{self.id}', False)
        
        self.write({
            'state': 'done',
            'lines_imported': imported,
        })
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
    
    def action_close(self):
        """Close wizard and return to sale order"""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'sale.order',
            'view_mode': 'form',
            'res_id': self.sale_order_id.id,
            'target': 'current',
        }
