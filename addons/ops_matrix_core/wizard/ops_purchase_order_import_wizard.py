# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
import logging
import json

_logger = logging.getLogger(__name__)

class OpsPurchaseOrderImportWizard(models.TransientModel):
    _name = 'ops.purchase.order.import.wizard'
    _description = 'Import Purchase Order Lines from Excel'
    
    purchase_order_id = fields.Many2one('purchase.order', 'Purchase Order', required=True)
    import_file = fields.Binary('Excel File', required=True)
    import_filename = fields.Char('Filename')
    template_file = fields.Binary('Template File', readonly=True, compute='_compute_template_file')
    template_filename = fields.Char('Template Filename', readonly=True, default='purchase_order_import_template.xlsx')
    
    state = fields.Selection([
        ('upload', 'Upload File'),
        ('validate', 'Validation Results'),
        ('done', 'Import Complete'),
    ], default='upload', string='State')
    
    validation_message = fields.Html('Validation Results', readonly=True)
    lines_imported = fields.Integer('Lines Imported', readonly=True)
    
    @api.depends('purchase_order_id')
    def _compute_template_file(self):
        for wizard in self:
            try:
                wizard.template_file = base64.b64encode(wizard._generate_template())
            except Exception as e:
                _logger.error(f"Error generating template: {e}")
                wizard.template_file = False
    
    def _generate_template(self):
        try:
            import xlsxwriter
        except ImportError:
            raise UserError("xlsxwriter Python library is not installed.")
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        
        instructions = workbook.add_worksheet('Instructions')
        bold = workbook.add_format({'bold': True, 'font_size': 12, 'font_color': '#1F4E78'})
        regular = workbook.add_format({'font_size': 10})
        
        instructions.write('A1', 'HOW TO USE THIS TEMPLATE', bold)
        instructions.write('A3', '1. Fill in the "Import Data" sheet with your products', regular)
        instructions.write('A4', '2. Product Code: Must match exactly an existing product code', regular)
        instructions.write('A5', '3. Quantity: Must be a positive number', regular)
        instructions.write('A6', '4. Unit Price: Optional - leave blank to use the product default cost', regular)
        instructions.write('A7', '5. Save the file and upload it in Odoo', regular)
        instructions.write('A8', '6. Click "Validate File" to check for errors before importing', regular)
        instructions.write('A9', '7. Do NOT modify column headers!', bold)
        instructions.set_column('A:A', 70)
        
        data_sheet = workbook.add_worksheet('Import Data')
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#366092',
            'font_color': 'white',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        headers = ['Product Code', 'Product Name (Reference Only)', 'Quantity', 'Unit Price (Optional)']
        for col, header in enumerate(headers):
            data_sheet.write(0, col, header, header_format)
        
        example_format = workbook.add_format({'italic': True, 'font_color': '#666666'})
        data_sheet.write(1, 0, 'PROD001', example_format)
        data_sheet.write(1, 1, 'Example Product Name', example_format)
        data_sheet.write(1, 2, 10, example_format)
        data_sheet.write(1, 3, 85.00, example_format)
        
        data_sheet.set_column('A:A', 20)
        data_sheet.set_column('B:B', 35)
        data_sheet.set_column('C:C', 12)
        data_sheet.set_column('D:D', 20)
        
        workbook.close()
        return output.getvalue()
    
    def action_download_template(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content?model={self._name}&id={self.id}&field=template_file&filename={self.template_filename}&download=true',
            'target': 'self',
        }
    
    def action_validate_import(self):
        self.ensure_one()
        if not self.import_file:
            raise UserError("Please upload an Excel file first.")
        
        errors = []
        warnings = []
        valid_lines = []
        
        try:
            from openpyxl import load_workbook
            file_data = base64.b64decode(self.import_file)
            workbook = load_workbook(io.BytesIO(file_data), read_only=True)
            
            if 'Import Data' not in workbook.sheetnames:
                raise UserError("Excel file must contain 'Import Data' sheet.")
            sheet = workbook['Import Data']
            
            expected_headers = ['Product Code', 'Product Name (Reference Only)', 'Quantity', 'Unit Price (Optional)']
            actual_headers = [str(sheet.cell(1, col+1).value or '').strip() for col in range(4)]
            
            if actual_headers != expected_headers:
                raise UserError(f"Column headers don't match template. Expected: {expected_headers}")
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:
                    continue
                
                product_code = str(row[0]).strip()
                quantity = row[2] if len(row) > 2 and row[2] else 0
                unit_price = row[3] if len(row) > 3 and row[3] else None
                
                product = self.env['product.product'].search([
                    ('default_code', '=', product_code)
                ], limit=1)
                
                if not product:
                    errors.append(f"Row {row_idx}: Product code '{product_code}' not found")
                    continue
                
                try:
                    qty = float(quantity)
                    if qty <= 0:
                        errors.append(f"Row {row_idx}: Quantity must be positive")
                        continue
                except (ValueError, TypeError):
                    errors.append(f"Row {row_idx}: Invalid quantity '{quantity}'")
                    continue
                
                if unit_price is not None:
                    try:
                        price = float(unit_price)
                        if price < 0:
                            errors.append(f"Row {row_idx}: Price cannot be negative")
                            continue
                    except (ValueError, TypeError):
                        warnings.append(f"Row {row_idx}: Invalid price, using default cost for {product.name}")
                        price = product.standard_price
                else:
                    price = product.standard_price
                
                valid_lines.append({
                    'product_id': product.id,
                    'product_qty': qty,
                    'price_unit': price,
                })
            
            msg = "<div style='padding: 15px; font-family: Arial, sans-serif;'>"
            if errors:
                msg += "<div style='background-color: #f8d7da; border-left: 4px solid #d9534f; padding: 10px; margin-bottom: 15px;'>"
                msg += "<h3 style='color: #d9534f; margin-top: 0;'>❌ Validation Failed</h3><ul>"
                for error in errors:
                    msg += f"<li>{error}</li>"
                msg += "</ul></div>"
            
            if warnings:
                msg += "<div style='background-color: #fff3cd; border-left: 4px solid #f0ad4e; padding: 10px; margin-bottom: 15px;'>"
                msg += "<h3 style='color: #856404; margin-top: 0;'>⚠️ Warnings</h3><ul>"
                for warning in warnings:
                    msg += f"<li>{warning}</li>"
                msg += "</ul></div>"
            
            if valid_lines and not errors:
                msg += "<div style='background-color: #d4edda; border-left: 4px solid #5cb85c; padding: 10px;'>"
                msg += "<h3 style='color: #5cb85c; margin-top: 0;'>✅ Validation Successful</h3>"
                msg += f"<p>Ready to import <strong>{len(valid_lines)} line(s)</strong> to purchase order <strong>{self.purchase_order_id.name}</strong></p>"
                msg += "</div>"
                
                self.env['ir.config_parameter'].sudo().set_param(
                    f'ops_po_import_lines_{self.id}',
                    json.dumps(valid_lines)
                )
            
            msg += "</div>"
            self.write({
                'validation_message': msg,
                'state': 'validate' if not errors else 'upload',
            })
            
        except Exception as e:
            _logger.exception("Error processing Excel file")
            raise UserError(f"Error processing file: {str(e)}")
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
    
    def action_import_lines(self):
        self.ensure_one()
        lines_str = self.env['ir.config_parameter'].sudo().get_param(f'ops_po_import_lines_{self.id}')
        if not lines_str:
            raise UserError("No validated lines found.")
        
        valid_lines = json.loads(lines_str)
        PurchaseOrderLine = self.env['purchase.order.line']
        imported = 0
        
        for line_data in valid_lines:
            product = self.env['product.product'].browse(line_data['product_id'])
            PurchaseOrderLine.create({
                'order_id': self.purchase_order_id.id,
                'product_id': product.id,
                'product_qty': line_data['product_qty'],
                'price_unit': line_data['price_unit'],
                'name': product.display_name,
                'date_planned': self.purchase_order_id.date_planned or fields.Datetime.now(),
                'product_uom': product.uom_id.id,
            })
            imported += 1
        
        self.env['ir.config_parameter'].sudo().set_param(f'ops_po_import_lines_{self.id}', False)
        self.write({'state': 'done', 'lines_imported': imported})
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
    
    def action_close(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'purchase.order',
            'view_mode': 'form',
            'res_id': self.purchase_order_id.id,
            'target': 'current',
        }
