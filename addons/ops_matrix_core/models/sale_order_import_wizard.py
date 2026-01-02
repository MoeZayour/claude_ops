from odoo import models, fields, api, _
from typing import TYPE_CHECKING
from odoo.exceptions import ValidationError
import base64
import io
try:
    import openpyxl
except ImportError:
    openpyxl = None

if TYPE_CHECKING:
    from odoo.api import Environment

class SaleOrderImportWizard(models.TransientModel):
    _name = 'sale.order.import.wizard'
    _description = 'Sales Order Excel Import Wizard'

    sale_order_id = fields.Many2one('sale.order', string='Sales Order', required=True)
    excel_file = fields.Binary(string='Excel File', required=True, help='Upload Excel file with columns: Section, Part Number, Quantity')
    file_name = fields.Char(string='File Name')
    import_result = fields.Text(string='Import Result', readonly=True)
    
    def action_validate_and_import(self):
        """Validate Excel file and import SO lines"""
        if not self.excel_file:
            raise ValidationError(_('Please upload an Excel file.'))
        
        if not openpyxl:
            raise ValidationError(_('openpyxl library is not installed. Please install it to use Excel import functionality.'))
        
        try:
            # Read Excel file
            excel_data = base64.b64decode(self.excel_file)
            workbook = openpyxl.load_workbook(io.BytesIO(excel_data))
            sheet = workbook.active
            
            # Get header row and validate columns
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)
            
            required_columns = ['Section', 'Part Number', 'Quantity']
            missing_columns = [col for col in required_columns if col not in headers]
            if missing_columns:
                raise ValidationError(_('Missing required columns: %s') % ', '.join(missing_columns))
            
            # Get column indices
            col_indices = {col: headers.index(col) for col in headers}
            
            # Validate all rows first (all-or-nothing approach)
            missing_parts = []
            product_model = self.env['product.product']
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue  # Skip empty rows
                    
                part_number = str(row[col_indices['Part Number']]).strip() if col_indices['Part Number'] < len(row) else ''
                if part_number and part_number.lower() not in ['none', 'nan']:
                    product = product_model.search([('default_code', '=', part_number)], limit=1)
                    if not product:
                        missing_parts.append(f"Row {row_idx}: {part_number} not found")
            
            # If any parts are missing, show error with all missing parts
            if missing_parts:
                error_message = _('The following part numbers were not found in the system:\n\n') + '\n'.join(missing_parts)
                self.import_result = error_message
                return {
                    'type': 'ir.actions.act_window',
                    'name': _('Import Validation Error'),
                    'res_model': 'sale.order.import.wizard',
                    'res_id': self.id,
                    'view_mode': 'form',
                    'target': 'new',
                }
            
            # All validations passed, create SO lines
            line_commands = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue  # Skip empty rows
                    
                section = str(row[col_indices['Section']]).strip() if col_indices['Section'] < len(row) else ''
                part_number = str(row[col_indices['Part Number']]).strip() if col_indices['Part Number'] < len(row) else ''
                
                # Handle quantity with proper ternary operator
                if col_indices['Quantity'] < len(row) and row[col_indices['Quantity']] not in [None, '']:
                    quantity = float(row[col_indices['Quantity']])
                else:
                    quantity = 0
                
                # Handle section rows
                if section and (not part_number or part_number.lower() in ['none', 'nan']):
                    line_commands.append([fields.Command.create({
                        'display_type': 'line_section',
                        'name': section,
                    })])
                elif part_number and quantity > 0:
                    # Find product
                    product = product_model.search([('default_code', '=', part_number)], limit=1)
                    if product:
                        # Get price from pricelist
                        pricelist = self.sale_order_id.pricelist_id
                        price = pricelist._get_product_price(product, quantity, partner=self.sale_order_id.partner_id)
                        
                        line_commands.append([fields.Command.create({
                            'product_id': product.id,
                            'product_uom_qty': quantity,
                            'price_unit': price,
                            'name': product.display_name,
                        })])
            
            # Add lines to sales order
            if line_commands:
                self.sale_order_id.write({
                    'order_line': line_commands
                })
            
            self.import_result = _('Successfully imported %d lines.') % len(line_commands)
            
            return {
                'type': 'ir.actions.act_window',
                'name': _('Import Successful'),
                'res_model': 'sale.order.import.wizard',
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'new',
            }
            
        except Exception as e:
            raise ValidationError(_('Error processing Excel file: %s') % str(e))
